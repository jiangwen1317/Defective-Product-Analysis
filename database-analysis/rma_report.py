"""
RMA 分析报告导出模块

生成 .xlsx 格式的 RMA 报告，包含 3 个 Sheet：
1. 设备概览 (Device Overview) - 主表摘要 + 核心指标
2. 详细指标 (Detailed Metrics) - KV 指标明细
3. 异常汇总 (Anomaly Summary) - Fail 记录 + 解析失败记录
"""
import logging
import os
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from database import DatabaseConnection, MetricsRepository

logger = logging.getLogger(__name__)

# 样式常量
_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class RMAReportGenerator:
    """RMA 报告生成器（.xlsx 格式）。"""

    def __init__(self, db: DatabaseConnection) -> None:
        self._db = db
        self._repo = MetricsRepository(db)

    def generate(
        self,
        output_path: str,
        *,
        device_name: Optional[str] = None,
        fw_version: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        include_sections: Optional[list[str]] = None,
        metric_keys: Optional[list[str]] = None,
    ) -> str:
        """生成 RMA 报告。

        Args:
            output_path: 输出文件路径（.xlsx）。
            device_name: 按设备名过滤。
            fw_version: 按固件版本过滤。
            date_from: 起始日期。
            date_to: 截止日期。
            include_sections: 包含的 Section 列表（默认全部）。
            metric_keys: 详细指标中展示的 metric_key 列表（默认全部）。

        Returns:
            生成的文件绝对路径。
        """
        with self._db.connect() as conn:
            # 获取主表数据
            summaries = self._repo.get_summaries(
                conn,
                device_name=device_name,
                fw_version=fw_version,
                date_from=date_from,
                date_to=date_to,
                limit=10000,
            )

            if not summaries:
                logger.warning("无数据可导出")

            # 获取指标数据
            all_metrics: list[dict] = []
            for s in summaries:
                metrics = self._repo.get_metrics(conn, summary_id=s["id"])
                if include_sections:
                    metrics = [m for m in metrics if m["section"] in include_sections]
                if metric_keys:
                    metrics = [m for m in metrics if m["metric_key"] in metric_keys]
                # 附加 summary 信息
                for m in metrics:
                    m["device_name"] = s.get("device_name")
                    m["fw_version"] = s.get("fw_version")
                    m["parsed_at"] = s.get("parsed_at")
                    m["overall_result"] = s.get("overall_result")
                all_metrics.extend(metrics)

            # 获取 Fail 记录
            fail_summaries = [s for s in summaries if s.get("overall_result") == "Fail"]

        # 创建工作簿
        wb = Workbook()

        # Sheet 1: 设备概览
        self._build_overview_sheet(wb, summaries)

        # Sheet 2: 详细指标
        self._build_metrics_sheet(wb, all_metrics)

        # Sheet 3: 异常汇总
        self._build_anomaly_sheet(wb, fail_summaries)

        # 保存
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        abs_path = os.path.abspath(output_path)
        logger.info("RMA 报告已生成: %s (%d 条记录, %d 条指标)", abs_path, len(summaries), len(all_metrics))
        return abs_path

    def _build_overview_sheet(self, wb: Workbook, summaries: list[dict]) -> None:
        """构建 Sheet1：设备概览。

        Args:
            wb: 工作簿对象。
            summaries: 主表摘要列表。
        """
        ws = wb.active
        ws.title = "设备概览"

        # 表头
        headers = [
            "序号", "设备名", "固件版本", "MP工具版本", "Flash ID",
            "原始坏块", "Cycles", "综合结果", "Fail Section",
            "WAI", "SLC PE Min", "SLC PE Max", "TLC PE Min", "TLC PE Max",
            "新增坏块", "解析状态", "解析时间",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER

        # 数据行
        for row_idx, s in enumerate(summaries, 2):
            values = [
                row_idx - 1,
                s.get("device_name", ""),
                s.get("fw_version", ""),
                s.get("mp_tool_version", ""),
                s.get("flash_id", ""),
                s.get("original_bad_block", ""),
                s.get("cycles", ""),
                s.get("overall_result", ""),
                s.get("fail_sections", ""),
                s.get("wai", ""),
                s.get("slc_pe_min", ""),
                s.get("slc_pe_max", ""),
                s.get("tlc_pe_min", ""),
                s.get("tlc_pe_max", ""),
                s.get("increase_bad_block", ""),
                s.get("parse_status", ""),
                s.get("parsed_at", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val if val is not None else "")
                cell.alignment = _CELL_ALIGN
                cell.border = _THIN_BORDER

            # 结果着色
            result_cell = ws.cell(row=row_idx, column=8)
            if s.get("overall_result") == "Fail":
                result_cell.fill = _FAIL_FILL
            elif s.get("overall_result") == "Pass":
                result_cell.fill = _PASS_FILL

        # 列宽自适应
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)].width = 18

    def _build_metrics_sheet(self, wb: Workbook, metrics: list[dict]) -> None:
        """构建 Sheet2：详细指标。

        Args:
            wb: 工作簿对象。
            metrics: 指标列表（含 summary 信息）。
        """
        ws = wb.create_sheet("详细指标")

        headers = [
            "设备名", "固件版本", "解析时间", "综合结果",
            "Section", "指标名(原始)", "指标名(清洗)",
            "原始值", "数值", "值类型", "前缀",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER

        for row_idx, m in enumerate(metrics, 2):
            values = [
                m.get("device_name", ""),
                m.get("fw_version", ""),
                m.get("parsed_at", ""),
                m.get("overall_result", ""),
                m.get("section", ""),
                m.get("metric_key_raw", ""),
                m.get("metric_key", ""),
                m.get("raw_value", ""),
                m.get("num_value", ""),
                m.get("value_type", ""),
                m.get("prefix", "") or "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val if val is not None else "")
                cell.alignment = _CELL_ALIGN
                cell.border = _THIN_BORDER

        # 列宽
        col_widths = [18, 30, 20, 10, 35, 25, 20, 20, 12, 10, 8]
        for col, width in enumerate(col_widths, 1):
            col_letter = chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)
            ws.column_dimensions[col_letter].width = width

    def _build_anomaly_sheet(self, wb: Workbook, fail_summaries: list[dict]) -> None:
        """构建 Sheet3：异常汇总。

        Args:
            wb: 工作簿对象。
            fail_summaries: Fail 的 summary 记录。
        """
        ws = wb.create_sheet("异常汇总")

        headers = [
            "序号", "设备名", "固件版本", "综合结果", "Fail Section",
            "WAI", "新增坏块", "解析状态", "解析时间",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER

        if not fail_summaries:
            ws.cell(row=2, column=1, value="无异常记录")
            return

        for row_idx, s in enumerate(fail_summaries, 2):
            values = [
                row_idx - 1,
                s.get("device_name", ""),
                s.get("fw_version", ""),
                s.get("overall_result", ""),
                s.get("fail_sections", ""),
                s.get("wai", ""),
                s.get("increase_bad_block", ""),
                s.get("parse_status", ""),
                s.get("parsed_at", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val if val is not None else "")
                cell.alignment = _CELL_ALIGN
                cell.border = _THIN_BORDER
                cell.fill = _FAIL_FILL

        # 列宽
        col_widths = [8, 18, 30, 10, 25, 12, 12, 10, 20]
        for col, width in enumerate(col_widths, 1):
            col_letter = chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)
            ws.column_dimensions[col_letter].width = width
