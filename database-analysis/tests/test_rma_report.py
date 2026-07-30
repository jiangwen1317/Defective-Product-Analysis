"""
rma_report 模块单元测试

覆盖 RMAReportGenerator 的报告生成逻辑：
Sheet 结构、数据行、Fail 汇总以及 Section/指标过滤。
"""
import json
import os

import pytest
from openpyxl import load_workbook
from rma_report import RMAReportGenerator


@pytest.fixture()
def seeded_db(tmp_db):
    """预置 1 条 Pass + 1 条 Fail 记录及其指标。"""
    db, repo = tmp_db
    with db.connect() as conn:
        sid1 = repo.insert_summary(
            conn,
            file_name="log1.txt",
            file_path="/tmp/log1.txt",
            file_size=100,
            file_mtime=1000.0,
            device_name="DM3720.026.04",
            fw_version="V3.2.9",
            overall_result="Pass",
            wai=161.136,
        )
        sid2 = repo.insert_summary(
            conn,
            file_name="log2.txt",
            file_path="/tmp/log2.txt",
            file_size=200,
            file_mtime=2000.0,
            device_name="DM3720.033.07",
            fw_version="V2.3.20",
            overall_result="Fail",
            fail_sections=json.dumps(["Wear_Detection"]),
            wai=149.2,
        )
        # 元组格式: (section, metric_key, metric_key_raw, raw_value,
        #            num_value, value_type, prefix, array_index)
        repo.insert_metrics_batch(conn, sid1, [
            ("header", "WAI", "WAI", "161.136", 161.136, "float", None, None),
        ])
        repo.insert_metrics_batch(conn, sid2, [
            ("header", "WAI", "WAI", "149.2", 149.2, "float", None, None),
            ("Wear_Detection", "wSLCMinPECycle", "wSLCMinPECycle",
             "0", 0.0, "decimal", None, None),
        ])
    return db, repo


class TestRMAReportGenerator:
    """测试报告生成的 Sheet 结构与数据内容。"""

    def test_generate_empty_db(self, tmp_db, tmp_path):
        """空数据库也应生成含 3 个 Sheet 的合法文件。"""
        db, _ = tmp_db
        output = str(tmp_path / "empty_report.xlsx")

        result_path = RMAReportGenerator(db).generate(output)

        wb = load_workbook(result_path)
        assert wb.sheetnames == ["设备概览", "详细指标", "异常汇总"]
        # 概览 Sheet 只有表头行
        assert wb["设备概览"].max_row == 1
        # 异常汇总显示占位文本
        assert wb["异常汇总"].cell(row=2, column=1).value == "无异常记录"

    def test_generate_returns_abs_path(self, tmp_db, tmp_path):
        db, _ = tmp_db
        output = str(tmp_path / "report.xlsx")
        result_path = RMAReportGenerator(db).generate(output)
        assert os.path.isabs(result_path)
        assert os.path.exists(result_path)

    def test_overview_sheet_rows(self, seeded_db, tmp_path):
        db, _ = seeded_db
        output = str(tmp_path / "report.xlsx")

        wb = load_workbook(RMAReportGenerator(db).generate(output))
        ws = wb["设备概览"]
        # 表头 + 2 条数据
        assert ws.max_row == 3
        devices = {ws.cell(row=r, column=2).value for r in (2, 3)}
        assert devices == {"DM3720.026.04", "DM3720.033.07"}

    def test_metrics_sheet_rows(self, seeded_db, tmp_path):
        db, _ = seeded_db
        output = str(tmp_path / "report.xlsx")

        wb = load_workbook(RMAReportGenerator(db).generate(output))
        ws = wb["详细指标"]
        # 表头 + 3 条指标
        assert ws.max_row == 4

    def test_anomaly_sheet_only_fail(self, seeded_db, tmp_path):
        db, _ = seeded_db
        output = str(tmp_path / "report.xlsx")

        wb = load_workbook(RMAReportGenerator(db).generate(output))
        ws = wb["异常汇总"]
        # 仅 1 条 Fail 记录
        assert ws.max_row == 2
        assert ws.cell(row=2, column=2).value == "DM3720.033.07"
        assert ws.cell(row=2, column=4).value == "Fail"

    def test_filter_by_device_name(self, seeded_db, tmp_path):
        db, _ = seeded_db
        output = str(tmp_path / "report.xlsx")

        wb = load_workbook(
            RMAReportGenerator(db).generate(output, device_name="DM3720.026.04")
        )
        ws = wb["设备概览"]
        assert ws.max_row == 2
        assert ws.cell(row=2, column=2).value == "DM3720.026.04"

    def test_filter_by_include_sections(self, seeded_db, tmp_path):
        """include_sections 应只保留指定 Section 的指标。"""
        db, _ = seeded_db
        output = str(tmp_path / "report.xlsx")

        wb = load_workbook(
            RMAReportGenerator(db).generate(
                output, include_sections=["Wear_Detection"]
            )
        )
        ws = wb["详细指标"]
        assert ws.max_row == 2
        assert ws.cell(row=2, column=5).value == "Wear_Detection"

    def test_filter_by_metric_keys(self, seeded_db, tmp_path):
        """metric_keys 应只保留指定指标。"""
        db, _ = seeded_db
        output = str(tmp_path / "report.xlsx")

        wb = load_workbook(
            RMAReportGenerator(db).generate(output, metric_keys=["WAI"])
        )
        ws = wb["详细指标"]
        # 两条记录各 1 个 WAI 指标
        assert ws.max_row == 3
        for r in (2, 3):
            assert ws.cell(row=r, column=7).value == "WAI"
