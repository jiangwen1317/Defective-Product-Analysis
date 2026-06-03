# -*- coding: utf-8 -*-
"""
RMA 报告导出模块

导出 Excel 格式的 RMA 分析报告。
"""

import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import Config, load_config
from .database import Database, get_database
from .query import QueryBuilder

logger = logging.getLogger(__name__)


class ReportExporter:
    """RMA 报告导出器"""

    # 样式定义
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    CELL_FONT = Font(size=10)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def __init__(
        self,
        db: Optional[Database] = None,
        config: Optional[Config] = None,
    ):
        """
        初始化导出器。

        Args:
            db: 数据库实例
            config: 配置对象
        """
        self.config = config or load_config()
        self.db = db or get_database()
        self.query = QueryBuilder(self.db)

    def _style_header(self, cell):
        """设置表头样式"""
        cell.fill = self.HEADER_FILL
        cell.font = self.HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = self.BORDER

    def _style_cell(self, cell, align: str = "left"):
        """设置单元格样式"""
        cell.font = self.CELL_FONT
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = self.BORDER

    def _style_result_cell(self, cell, value: str):
        """根据结果值设置单元格样式"""
        self._style_cell(cell, "center")
        if value.lower() == "pass":
            cell.fill = self.PASS_FILL
        elif value.lower() == "fail":
            cell.fill = self.FAIL_FILL

    def _auto_column_width(self, worksheet, min_width: int = 10, max_width: int = 50):
        """自动调整列宽"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def export_device_overview(self, worksheet):
        """
        Sheet1: 设备概览

        包含: 设备信息 + 核心指标
        """
        # 获取设备概览数据
        summaries = self.query.get_all_summaries()

        # 写入表头
        headers = [
            "ID", "设备名称", "固件版本", "工具版本",
            "测试循环", "测试结果", "WAI", "WA(TLC)", "WA(SLC)",
            "解析状态", "文件名", "创建时间",
        ]
        worksheet.append(headers)

        for col_idx, header in enumerate(headers, 1):
            self._style_header(worksheet.cell(row=1, column=col_idx))

        # 写入数据
        for row_idx, summary in enumerate(summaries, 2):
            # 获取核心指标
            metrics = self.query.get_metrics_by_summary(summary.id)

            # 提取关键指标值
            wai = next(
                (m.num_value for m in metrics if m.metric_key == "WAI"),
                None
            )
            wa_tlc = next(
                (m.num_value for m in metrics if m.metric_key == "WA(TLC)"),
                None
            )
            wa_slc = next(
                (m.num_value for m in metrics if m.metric_key == "WA(SLC)"),
                None
            )

            row_data = [
                summary.id,
                summary.device_name,
                summary.fw_version,
                summary.tool_version,
                summary.test_cycles,
                summary.test_result,
                wai,
                wa_tlc,
                wa_slc,
                summary.parse_status,
                summary.file_name,
                datetime.fromtimestamp(summary.created_at).strftime("%Y-%m-%d %H:%M:%S"),
            ]
            worksheet.append(row_data)

            # 设置数据行样式
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                self._style_cell(cell)

                # 测试结果着色
                if col_idx == 6:  # test_result 列
                    self._style_result_cell(cell, value)

                # 解析状态着色
                if col_idx == 10:  # parse_status 列
                    if value == "Failed":
                        cell.fill = self.FAIL_FILL
                    elif value == "Pending":
                        cell.fill = self.WARNING_FILL

        # 自动调整列宽
        self._auto_column_width(worksheet)

        logger.info(f"Sheet1 设备概览: 已写入 {len(summaries)} 条记录")

    def export_detailed_metrics(
        self,
        worksheet,
        metric_keys: Optional[List[str]] = None,
    ):
        """
        Sheet2: 详细指标 (动态透视的 KV 宽表)

        Args:
            worksheet: Excel 工作表
            metric_keys: 要展示的指标键名列表 (为空则展示所有)
        """
        # 获取所有摘要
        summaries = self.query.get_all_summaries(parse_status="Success")

        if not summaries:
            worksheet.append(["无数据"])
            logger.warning("Sheet2: 无解析成功的记录")
            return

        # 获取指标键名 (如果不指定,则获取前 20 个)
        if not metric_keys:
            all_keys = self.query.get_unique_metric_keys()
            metric_keys = all_keys[:20]  # 限制列数

        # 构建表头
        headers = ["设备名称", "固件版本", "测试循环", "结果"] + metric_keys
        worksheet.append(headers)

        for col_idx, header in enumerate(headers, 1):
            self._style_header(worksheet.cell(row=1, column=col_idx))

        # 批量获取指标数据
        summary_ids = [s.id for s in summaries]
        metrics_data = self.query.get_metric_values(summary_ids, metric_keys)

        # 写入数据
        for row_idx, summary in enumerate(summaries, 2):
            device_metrics = metrics_data.get(summary.id, {})

            row_data = [
                summary.device_name,
                summary.fw_version,
                summary.test_cycles,
                summary.test_result,
            ]

            for key in metric_keys:
                key_data = device_metrics.get(key, {})
                value = key_data.get("raw") or key_data.get("num") or ""
                row_data.append(value)

            worksheet.append(row_data)

            # 设置数据行样式
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                self._style_cell(cell)

                # 测试结果着色
                if col_idx == 4:
                    self._style_result_cell(cell, value)

        # 自动调整列宽
        self._auto_column_width(worksheet, max_width=30)

        logger.info(f"Sheet2 详细指标: 已写入 {len(summaries)} 条记录, {len(metric_keys)} 个指标")

    def export_anomaly_summary(self, worksheet):
        """
        Sheet3: 异常汇总

        包含: parse_status='Failed' 的记录及错误原因
        """
        # 获取失败记录
        failed_records = self.query.get_failed_summary()

        # 写入表头
        headers = [
            "ID", "设备名称", "固件版本", "文件名",
            "解析状态", "测试结果", "错误信息", "创建时间",
        ]
        worksheet.append(headers)

        for col_idx, header in enumerate(headers, 1):
            self._style_header(worksheet.cell(row=1, column=col_idx))

        # 写入数据
        for row_idx, record in enumerate(failed_records, 2):
            row_data = [
                record["id"],
                record["device_name"],
                record["fw_version"],
                record["file_name"],
                record["parse_status"],
                record["test_result"] or "N/A",
                record["parse_error"] or "N/A",
                datetime.fromtimestamp(record["created_at"]).strftime("%Y-%m-%d %H:%M:%S"),
            ]
            worksheet.append(row_data)

            # 设置数据行样式
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                self._style_cell(cell)

                # 异常类型着色
                if col_idx == 5 and value == "Failed":  # parse_status 列
                    cell.fill = self.FAIL_FILL

        # 自动调整列宽
        self._auto_column_width(worksheet)

        logger.info(f"Sheet3 异常汇总: 已写入 {len(failed_records)} 条记录")

    def export(
        self,
        output_path: str,
        metric_keys: Optional[List[str]] = None,
    ) -> str:
        """
        导出 RMA 报告。

        Args:
            output_path: 输出文件路径
            metric_keys: 要展示的指标键名列表

        Returns:
            输出文件路径
        """
        logger.info(f"开始导出 RMA 报告: {output_path}")

        # 确保输出目录存在
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        # 创建工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认工作表

        # Sheet1: 设备概览
        ws1 = wb.create_sheet("设备概览")
        self.export_device_overview(ws1)

        # Sheet2: 详细指标
        ws2 = wb.create_sheet("详细指标")
        self.export_detailed_metrics(ws2, metric_keys)

        # Sheet3: 异常汇总
        ws3 = wb.create_sheet("异常汇总")
        self.export_anomaly_summary(ws3)

        # 保存文件
        wb.save(output_path)
        logger.info(f"RMA 报告已导出: {output_path}")

        return output_path

    def export_default(
        self,
        output_dir: Optional[str] = None,
        filename: Optional[str] = None,
    ) -> str:
        """
        使用默认配置导出报告。

        Args:
            output_dir: 输出目录
            filename: 文件名 (不含扩展名)

        Returns:
            输出文件路径
        """
        if output_dir is None:
            output_dir = self.config.export.output_dir

        os.makedirs(output_dir, exist_ok=True)

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"RMA_Report_{timestamp}"

        output_path = os.path.join(output_dir, f"{filename}.xlsx")
        return self.export(output_path)


def export_report(
    output_path: Optional[str] = None,
    db: Optional[Database] = None,
) -> str:
    """
    便捷函数: 导出 RMA 报告。

    Args:
        output_path: 输出文件路径
        db: 数据库实例

    Returns:
        输出文件路径
    """
    exporter = ReportExporter(db)
    if output_path:
        return exporter.export(output_path)
    else:
        return exporter.export_default()
